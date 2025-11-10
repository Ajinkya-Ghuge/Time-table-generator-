# app.py
from flask import Flask, render_template, send_file, request, jsonify, flash, redirect, url_for, session
import os
import io
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import tempfile
from werkzeug.utils import secure_filename
from collections import defaultdict
app = Flask(__name__)
app.secret_key = 'super_secret_key'  # For flash messages and session
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
# Create upload folder if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Original hardcoded data as fallback
def get_default_data():
    # --- Course Data ---
    course_data = [
        # Year 2: 6 main subjects + 2 NEP-like (total 8)
        {'course_id': '201', 'name': 'Data Structures', 'program': 'B.Tech CSE', 'year': 2, 'divisions': 'A,B,C', 'credit': 5, 'hours_per_week': 5, 'theory_hours': 3, 'practical_hours': 2, 'lab_required': 'Yes', 'room_type': 'Lab', 'central': 'No', 'faculty_id': 'F001', 'batch_size': 60, 'preferred_slot': None},
        {'course_id': '202', 'name': 'Digital Logic Design', 'program': 'B.Tech CSE', 'year': 2, 'divisions': 'A,B,C', 'credit': 5, 'hours_per_week': 5, 'theory_hours': 3, 'practical_hours': 2, 'lab_required': 'Yes', 'room_type': 'Lab', 'central': 'No', 'faculty_id': 'F002', 'batch_size': 60, 'preferred_slot': None},
        {'course_id': '203', 'name': 'Mathematics III', 'program': 'B.Tech CSE', 'year': 2, 'divisions': 'A,B,C', 'credit': 4, 'hours_per_week': 4, 'theory_hours': 4, 'practical_hours': 0, 'lab_required': 'No', 'room_type': 'Classroom', 'central': 'No', 'faculty_id': 'F003', 'batch_size': 60, 'preferred_slot': None},
        {'course_id': '204', 'name': 'Computer Organization', 'program': 'B.Tech CSE', 'year': 2, 'divisions': 'A,B,C', 'credit': 5, 'hours_per_week': 5, 'theory_hours': 3, 'practical_hours': 2, 'lab_required': 'Yes', 'room_type': 'Lab', 'central': 'No', 'faculty_id': 'F013', 'batch_size': 60, 'preferred_slot': None},
        {'course_id': '205', 'name': 'Programming in Java', 'program': 'B.Tech CSE', 'year': 2, 'divisions': 'A,B,C', 'credit': 5, 'hours_per_week': 5, 'theory_hours': 3, 'practical_hours': 2, 'lab_required': 'Yes', 'room_type': 'Lab', 'central': 'No', 'faculty_id': 'F014', 'batch_size': 60, 'preferred_slot': None},
        {'course_id': '206', 'name': 'Discrete Mathematics', 'program': 'B.Tech CSE', 'year': 2, 'divisions': 'A,B,C', 'credit': 4, 'hours_per_week': 4, 'theory_hours': 4, 'practical_hours': 0, 'lab_required': 'No', 'room_type': 'Classroom', 'central': 'No', 'faculty_id': 'F015', 'batch_size': 60, 'preferred_slot': None},
        {'course_id': '207', 'name': 'OE - Environmental Studies (NEP)', 'program': 'B.Tech CSE', 'year': 2, 'divisions': 'A,B,C', 'credit': 4, 'hours_per_week': 4, 'theory_hours': 4, 'practical_hours': 0, 'lab_required': 'No', 'room_type': 'Classroom', 'central': 'Yes', 'faculty_id': 'F010', 'batch_size': 180, 'preferred_slot': 'Wed-2'},
        {'course_id': '208', 'name': 'Communication Skills (NEP)', 'program': 'B.Tech CSE', 'year': 2, 'divisions': 'A,B,C', 'credit': 4, 'hours_per_week': 4, 'theory_hours': 4, 'practical_hours': 0, 'lab_required': 'No', 'room_type': 'Classroom', 'central': 'Yes', 'faculty_id': 'F016', 'batch_size': 180, 'preferred_slot': None},
       
        # Year 3: 6 main subjects + 2 NEP-like (total 8)
        {'course_id': '301', 'name': 'Operating Systems', 'program': 'B.Tech CSE', 'year': 3, 'divisions': 'A,B,C', 'credit': 5, 'hours_per_week': 5, 'theory_hours': 3, 'practical_hours': 2, 'lab_required': 'Yes', 'room_type': 'Lab', 'central': 'No', 'faculty_id': 'F004', 'batch_size': 60, 'preferred_slot': None},
        {'course_id': '302', 'name': 'Database Systems', 'program': 'B.Tech CSE', 'year': 3, 'divisions': 'A,B,C', 'credit': 5, 'hours_per_week': 5, 'theory_hours': 3, 'practical_hours': 2, 'lab_required': 'Yes', 'room_type': 'Lab', 'central': 'No', 'faculty_id': 'F005', 'batch_size': 60, 'preferred_slot': None},
        {'course_id': '303', 'name': 'Computer Networks', 'program': 'B.Tech CSE', 'year': 3, 'divisions': 'A,B,C', 'credit': 5, 'hours_per_week': 5, 'theory_hours': 3, 'practical_hours': 2, 'lab_required': 'Yes', 'room_type': 'Lab', 'central': 'No', 'faculty_id': 'F006', 'batch_size': 60, 'preferred_slot': None},
        {'course_id': '304', 'name': 'Software Engineering', 'program': 'B.Tech CSE', 'year': 3, 'divisions': 'A,B,C', 'credit': 5, 'hours_per_week': 5, 'theory_hours': 3, 'practical_hours': 2, 'lab_required': 'Yes', 'room_type': 'Lab', 'central': 'No', 'faculty_id': 'F017', 'batch_size': 60, 'preferred_slot': None},
        {'course_id': '305', 'name': 'Algorithms', 'program': 'B.Tech CSE', 'year': 3, 'divisions': 'A,B,C', 'credit': 4, 'hours_per_week': 4, 'theory_hours': 4, 'practical_hours': 0, 'lab_required': 'No', 'room_type': 'Classroom', 'central': 'No', 'faculty_id': 'F018', 'batch_size': 60, 'preferred_slot': None},
        {'course_id': '306', 'name': 'Web Development', 'program': 'B.Tech CSE', 'year': 3, 'divisions': 'A,B,C', 'credit': 5, 'hours_per_week': 5, 'theory_hours': 3, 'practical_hours': 2, 'lab_required': 'Yes', 'room_type': 'Lab', 'central': 'No', 'faculty_id': 'F019', 'batch_size': 60, 'preferred_slot': None},
        {'course_id': '307', 'name': 'OE - Constitution of India (NEP)', 'program': 'B.Tech CSE', 'year': 3, 'divisions': 'A,B,C', 'credit': 4, 'hours_per_week': 4, 'theory_hours': 4, 'practical_hours': 0, 'lab_required': 'No', 'room_type': 'Classroom', 'central': 'Yes', 'faculty_id': 'F011', 'batch_size': 180, 'preferred_slot': 'Thu-3'},
        {'course_id': '308', 'name': 'Probability and Statistics (NEP)', 'program': 'B.Tech CSE', 'year': 3, 'divisions': 'A,B,C', 'credit': 4, 'hours_per_week': 4, 'theory_hours': 4, 'practical_hours': 0, 'lab_required': 'No', 'room_type': 'Classroom', 'central': 'Yes', 'faculty_id': 'F020', 'batch_size': 180, 'preferred_slot': None},
       
        # Year 4: 6 main subjects + 2 NEP-like (total 8)
        {'course_id': '401', 'name': 'Machine Learning', 'program': 'B.Tech CSE', 'year': 4, 'divisions': 'A,B,C', 'credit': 5, 'hours_per_week': 5, 'theory_hours': 3, 'practical_hours': 2, 'lab_required': 'Yes', 'room_type': 'Lab', 'central': 'No', 'faculty_id': 'F007', 'batch_size': 60, 'preferred_slot': None},
        {'course_id': '402', 'name': 'Cloud Computing', 'program': 'B.Tech CSE', 'year': 4, 'divisions': 'A,B,C', 'credit': 5, 'hours_per_week': 5, 'theory_hours': 3, 'practical_hours': 2, 'lab_required': 'Yes', 'room_type': 'Lab', 'central': 'No', 'faculty_id': 'F008', 'batch_size': 60, 'preferred_slot': None},
        {'course_id': '403', 'name': 'Major Project', 'program': 'B.Tech CSE', 'year': 4, 'divisions': 'A,B,C', 'credit': 8, 'hours_per_week': 8, 'theory_hours': 0, 'practical_hours': 8, 'lab_required': 'Yes', 'room_type': 'Lab', 'central': 'No', 'faculty_id': 'F009', 'batch_size': 60, 'preferred_slot': 'Fri-2,3'},
        {'course_id': '404', 'name': 'Artificial Intelligence', 'program': 'B.Tech CSE', 'year': 4, 'divisions': 'A,B,C', 'credit': 5, 'hours_per_week': 5, 'theory_hours': 3, 'practical_hours': 2, 'lab_required': 'Yes', 'room_type': 'Lab', 'central': 'No', 'faculty_id': 'F021', 'batch_size': 60, 'preferred_slot': None},
        {'course_id': '405', 'name': 'Big Data Analytics', 'program': 'B.Tech CSE', 'year': 4, 'divisions': 'A,B,C', 'credit': 5, 'hours_per_week': 5, 'theory_hours': 3, 'practical_hours': 2, 'lab_required': 'Yes', 'room_type': 'Lab', 'central': 'No', 'faculty_id': 'F022', 'batch_size': 60, 'preferred_slot': None},
        {'course_id': '406', 'name': 'Cyber Security', 'program': 'B.Tech CSE', 'year': 4, 'divisions': 'A,B,C', 'credit': 4, 'hours_per_week': 4, 'theory_hours': 4, 'practical_hours': 0, 'lab_required': 'No', 'room_type': 'Classroom', 'central': 'No', 'faculty_id': 'F023', 'batch_size': 60, 'preferred_slot': None},
        {'course_id': '407', 'name': 'MDM - Seminar (NEP)', 'program': 'B.Tech CSE', 'year': 4, 'divisions': 'A,B,C', 'credit': 4, 'hours_per_week': 4, 'theory_hours': 4, 'practical_hours': 0, 'lab_required': 'No', 'room_type': 'Seminar', 'central': 'Yes', 'faculty_id': 'F012', 'batch_size': 180, 'preferred_slot': 'Sat-1'},
        {'course_id': '408', 'name': 'Multi Disciplinary Minor (NEP)', 'program': 'B.Tech CSE', 'year': 4, 'divisions': 'A,B,C', 'credit': 4, 'hours_per_week': 4, 'theory_hours': 4, 'practical_hours': 0, 'lab_required': 'No', 'room_type': 'Classroom', 'central': 'Yes', 'faculty_id': 'F024', 'batch_size': 180, 'preferred_slot': None},
    ]
    dfCourse = pd.DataFrame(course_data)
    # --- Faculty Data ---
    faculty_data = [
        {'faculty_id': 'F001', 'name': 'Dr. A. Sharma', 'expertise': '[201] (Data Structures)', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[Mon-2]', 'consecutive_limit': 6},
        {'faculty_id': 'F002', 'name': 'Prof. B. Patil', 'expertise': '[202] (Digital Logic Design)', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[]', 'consecutive_limit': 6},
        {'faculty_id': 'F003', 'name': 'Dr. C. Mehta', 'expertise': '[203] (Mathematics III)', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[Tue-2]', 'consecutive_limit': 6},
        {'faculty_id': 'F004', 'name': 'Prof. D. Kulkarni', 'expertise': '[301] (Operating Systems)', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[]', 'consecutive_limit': 6},
        {'faculty_id': 'F005', 'name': 'Dr. E. Iyer', 'expertise': '[302] (Database Systems)', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[Fri-3]', 'consecutive_limit': 6},
        {'faculty_id': 'F006', 'name': 'Dr. F. Singh', 'expertise': '[303] (Computer Networks)', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[]', 'consecutive_limit': 6},
        {'faculty_id': 'F007', 'name': 'Prof. G. Deshmukh', 'expertise': '[401] (Machine Learning)', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[Wed-1]', 'consecutive_limit': 6},
        {'faculty_id': 'F008', 'name': 'Dr. H. Verma', 'expertise': '[402] (Cloud Computing)', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[]', 'consecutive_limit': 6},
        {'faculty_id': 'F009', 'name': 'Prof. I. Rao', 'expertise': '[403] (Major Project)', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[Fri-2, Fri-3]', 'consecutive_limit': 6},
        {'faculty_id': 'F010', 'name': 'Dr. J. Joshi', 'expertise': '[207] (OE - Environmental Studies (NEP))', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[Wed-2]', 'consecutive_limit': 6},
        {'faculty_id': 'F011', 'name': 'Prof. K. Banerjee', 'expertise': '[307] (OE - Constitution of India (NEP))', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[Thu-3]', 'consecutive_limit': 6},
        {'faculty_id': 'F012', 'name': 'Dr. L. Reddy', 'expertise': '[407] (MDM - Seminar (NEP))', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[Sat-1]', 'consecutive_limit': 6},
        {'faculty_id': 'F013', 'name': 'Dr. M. Khan', 'expertise': '[204] (Computer Organization)', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[]', 'consecutive_limit': 6},
        {'faculty_id': 'F014', 'name': 'Prof. N. Singh', 'expertise': '[205] (Programming in Java)', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[]', 'consecutive_limit': 6},
        {'faculty_id': 'F015', 'name': 'Dr. O. Patel', 'expertise': '[206] (Discrete Mathematics)', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[]', 'consecutive_limit': 6},
        {'faculty_id': 'F016', 'name': 'Prof. P. Lee', 'expertise': '[208] (Communication Skills (NEP))', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[]', 'consecutive_limit': 6},
        {'faculty_id': 'F017', 'name': 'Dr. Q. Gomez', 'expertise': '[304] (Software Engineering)', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[]', 'consecutive_limit': 6},
        {'faculty_id': 'F018', 'name': 'Prof. R. Kim', 'expertise': '[305] (Algorithms)', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[]', 'consecutive_limit': 6},
        {'faculty_id': 'F019', 'name': 'Dr. S. Wong', 'expertise': '[306] (Web Development)', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[]', 'consecutive_limit': 6},
        {'faculty_id': 'F020', 'name': 'Prof. T. Chen', 'expertise': '[308] (Probability and Statistics (NEP))', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[]', 'consecutive_limit': 6},
        {'faculty_id': 'F021', 'name': 'Dr. U. Lopez', 'expertise': '[404] (Artificial Intelligence)', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[]', 'consecutive_limit': 6},
        {'faculty_id': 'F022', 'name': 'Prof. V. Garcia', 'expertise': '[405] (Big Data Analytics)', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[]', 'consecutive_limit': 6},
        {'faculty_id': 'F023', 'name': 'Dr. W. Martinez', 'expertise': '[406] (Cyber Security)', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[]', 'consecutive_limit': 6},
        {'faculty_id': 'F024', 'name': 'Prof. X. Rodriguez', 'expertise': '[408] (Multi Disciplinary Minor (NEP))', 'max_hours_per_week': 40, 'unavailable_slots': '[]', 'preferred_slots': '[]', 'consecutive_limit': 6},
    ]
    dfFaculty = pd.DataFrame(faculty_data)
    # --- Room Data ---
    room_data = [
        {'room_id': 'R101', 'capacity': 60, 'type': 'theory', 'available_slots': None},
        {'room_id': 'R102', 'capacity': 60, 'type': 'theory', 'available_slots': None},
        {'room_id': 'R201', 'capacity': 30, 'type': 'lab', 'available_slots': None},
        {'room_id': 'R202', 'capacity': 30, 'type': 'lab', 'available_slots': None},
        {'room_id': 'CL1', 'capacity': 40, 'type': 'computer_lab', 'available_slots': None},
        {'room_id': 'CL2', 'capacity': 40, 'type': 'computer_lab', 'available_slots': None},
        {'room_id': 'AUD1', 'capacity': 120, 'type': 'theory', 'available_slots': None},
        {'room_id': 'R301', 'capacity': 25, 'type': 'theory', 'available_slots': None},
        {'room_id': 'AUD2', 'capacity': 240, 'type': 'theory', 'available_slots': None},
        {'room_id': 'AUD3', 'capacity': 240, 'type': 'theory', 'available_slots': None},
        {'room_id': 'AUD4', 'capacity': 240, 'type': 'theory', 'available_slots': None},
        {'room_id': 'LAB_BIG', 'capacity': 240, 'type': 'lab', 'available_slots': None},
        {'room_id': 'LAB2', 'capacity': 240, 'type': 'lab', 'available_slots': None},
        {'room_id': 'LAB3', 'capacity': 240, 'type': 'lab', 'available_slots': None},
    ]
    dfRoom = pd.DataFrame(room_data)
    return dfCourse, dfFaculty, dfRoom

def parse_list(s):
    if not s or s == '[]':
        return []
    if isinstance(s, list):
        return s  # Already parsed
    s = str(s).strip('[]')
    return [item.strip() for item in s.split(',')]

def preprocess_data(dfCourse, dfFaculty, dfRoom):
    # Preprocess divisions for courses
    dfCourse['divisions_list'] = dfCourse['divisions'].apply(lambda x: x.split(',') if pd.notna(x) else [])
    dfCourse['divisions_num'] = dfCourse['divisions_list'].apply(len)
   
    # Preprocess faculty slots (safe apply)
    if 'unavailable_slots' in dfFaculty.columns:
        dfFaculty['unavailable_slots'] = dfFaculty['unavailable_slots'].apply(parse_list)
    if 'preferred_slots' in dfFaculty.columns:
        dfFaculty['preferred_slots'] = dfFaculty['preferred_slots'].apply(parse_list)
   
    return dfCourse, dfFaculty, dfRoom

def generate_timetables(dfCourse, dfFaculty, dfRoom):
    dfCourse, dfFaculty, dfRoom = preprocess_data(dfCourse, dfFaculty, dfRoom)
   
    # --- Constants ---
    COLLEGE_START = 10*60 + 15  # 10:15 AM
    COLLEGE_END = 17*60 + 30  # 5:30 PM
    THEORY_DURATION = 60  # 1 hour
    WORKING_DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    BREAKS = [
        {"name": "Lunch Break", "start": 12*60 + 15, "duration": 60},  # 12:15–13:15
        {"name": "Tea Break", "start": 15*60 + 15, "duration": 15},  # 3:15–3:30
    ]
    def minutes_to_time(m):
        hour = m // 60
        minute = m % 60
        am_pm = 'am' if hour < 12 else 'pm'
        hour12 = hour if hour <= 12 else hour - 12
        if hour12 == 0:
            hour12 = 12
        return f"{hour12}:{minute:02d} {am_pm}"
    def generate_slots():
        slots = []
        t = COLLEGE_START
        while t < COLLEGE_END:
            brk = next((br for br in BREAKS if br["start"] == t), None)
            if brk:
                br_start = minutes_to_time(brk["start"])
                br_end = minutes_to_time(brk["start"] + brk["duration"])
                slots.append(f"{br_start}-{br_end}")
                t += brk["duration"]
            else:
                start = minutes_to_time(t)
                end = minutes_to_time(t + THEORY_DURATION)
                slots.append(f"{start}-{end}")
                t += THEORY_DURATION
        return slots
    def create_all_timetables(dfCourse):
        timetables = {}
        slots = generate_slots()
        grouped = dfCourse.groupby("year")["divisions_num"].max()
        div_map = 'ABC'
        for year, max_divs in grouped.items():
            for div in range(1, max_divs + 1):
                key = f"Year{year}_Div{div_map[div - 1]}"
                df = pd.DataFrame(
                    [["" for _ in slots] for _ in WORKING_DAYS],
                    index=WORKING_DAYS,
                    columns=slots
                )
                timetables[key] = df
        return timetables
    timetables = create_all_timetables(dfCourse)
    # --- Slot map ---
    slots = generate_slots()
    lecture_slot_times = [s for s in slots if not s.startswith('12:15') and not s.startswith('15:15')]
    slot_number_map = {}
    for day in WORKING_DAYS:
        for slot_num, slot_time in enumerate(lecture_slot_times, 1):
            key = f"{day}-{slot_num}"
            slot_number_map[key] = slot_time
    # --- Prepare sessions ---
    sessions = []
    for idx, row in dfCourse.iterrows():
        year = row['year']
        divs = row['divisions_list']
        faculty = row['faculty_id']
        name = row['name']
        central = row['central'] == 'Yes'
        theory_h = row['theory_hours']
        prac_h = row['practical_hours']
        preferred = row['preferred_slot']
        batch_size = row['batch_size']
        preferred_list = [p.strip() for p in preferred.split(',')] if preferred else []
        if central:
            # For theory
            for i in range(theory_h):
                sess_pref = preferred_list[i] if i < len(preferred_list) else None
                sessions.append({'type': 'theory', 'course_id': row['course_id'], 'name': name, 'faculty': faculty, 'year': year, 'divs': divs, 'central': True, 'room_type': 'theory', 'batch_size': batch_size, 'preferred': sess_pref, 'duration': 1})
            # For practical
            for i in range(prac_h // 2):
                sess_pref = preferred_list[i + theory_h] if i + theory_h < len(preferred_list) else None
                sessions.append({'type': 'practical', 'course_id': row['course_id'], 'name': name, 'faculty': faculty, 'year': year, 'divs': divs, 'central': True, 'room_type': 'lab', 'batch_size': batch_size, 'preferred': sess_pref, 'duration': 2})
        else:
            for div in divs:
                pref_idx = 0
                for i in range(theory_h):
                    sess_pref = preferred_list[pref_idx] if pref_idx < len(preferred_list) else None
                    pref_idx += 1
                    sessions.append({'type': 'theory', 'course_id': row['course_id'], 'name': name, 'faculty': faculty, 'year': year, 'divs': [div], 'central': False, 'room_type': 'theory', 'batch_size': 60, 'preferred': sess_pref, 'duration': 1})
                for i in range(prac_h // 2):
                    sess_pref = preferred_list[pref_idx] if pref_idx < len(preferred_list) else None
                    pref_idx += 1
                    sessions.append({'type': 'practical', 'course_id': row['course_id'], 'name': name, 'faculty': faculty, 'year': year, 'divs': [div], 'central': False, 'room_type': 'lab', 'batch_size': 60, 'preferred': sess_pref, 'duration': 2})
    # --- Sort sessions ---
    sessions = sorted(sessions, key=lambda s: s['preferred'] is None)
    # --- Scheduling structures ---
    faculty_schedule = defaultdict(list)
    room_schedule = defaultdict(list)
    faculty_hours = defaultdict(int)
    # --- Schedule sessions ---
    for session in sessions:
        possible_slots = []
        faculty = session['faculty']
        max_hours = dfFaculty[dfFaculty['faculty_id'] == faculty]['max_hours_per_week'].item()
        consecutive_limit = dfFaculty[dfFaculty['faculty_id'] == faculty]['consecutive_limit'].item()
        unavailable = dfFaculty[dfFaculty['faculty_id'] == faculty]['unavailable_slots'].item()
        room_type = session['room_type']
        duration = session['duration']
        if room_type == 'Seminar':
            room_type = 'theory'
        for day in WORKING_DAYS:
            day_hours = len([s for d, s in faculty_schedule[faculty] if d == day])
            if day_hours + duration > 6:
                continue
            for slot_num in range(1, 7):
                if duration == 2 and slot_num not in [1,3,5]:
                    continue
                key = f"{day}-{slot_num}"
                slot_time = slot_number_map[key]
                slot_time2 = None
                key2 = None
                if duration == 2:
                    slot_num2 = slot_num + 1
                    key2 = f"{day}-{slot_num2}"
                    slot_time2 = slot_number_map[key2]
                if session['preferred'] and session['preferred'] != key:
                    continue
                if key in unavailable or (duration == 2 and key2 in unavailable):
                    continue
                if any(d == day and s in [slot_time, slot_time2] for d, s in faculty_schedule[faculty]):
                    continue
                if faculty_hours[faculty] + duration > max_hours:
                    continue
                day_slots = [s for d, s in faculty_schedule[faculty] if d == day]
                day_indices = sorted([lecture_slot_times.index(s) for s in day_slots])
                new_index = lecture_slot_times.index(slot_time)
                day_indices.append(new_index)
                if duration == 2:
                    new_index2 = new_index + 1
                    day_indices.append(new_index2)
                day_indices = sorted(day_indices)
                curr_con = 1
                max_con = 1
                for j in range(1, len(day_indices)):
                    if day_indices[j] == day_indices[j-1] + 1:
                        curr_con += 1
                        max_con = max(max_con, curr_con)
                    else:
                        curr_con = 1
                if max_con > consecutive_limit:
                    continue
                possible_rooms = dfRoom[dfRoom['type'] == 'theory' if room_type == 'theory' else dfRoom['type'].str.contains('lab')]
                room_found = False
                chosen_room = None
                for r_idx, r_row in possible_rooms.iterrows():
                    r_id = r_row['room_id']
                    if any(d == day and s in [slot_time, slot_time2] for d, s in room_schedule[r_id]):
                        continue
                    chosen_room = r_id
                    room_found = True
                    break
                if not room_found:
                    continue
                empty = True
                for div in session['divs']:
                    key_tt = f"Year{session['year']}_Div{div}"
                    if timetables[key_tt].at[day, slot_time] != "" or (duration == 2 and timetables[key_tt].at[day, slot_time2] != ""):
                        empty = False
                        break
                if not empty:
                    continue
                # Check for only 1 practical per day per division
                if session['type'] == 'practical':
                    has_practical = False
                    for div in session['divs']:
                        key_tt = f"Year{session['year']}_Div{div}"
                        day_schedule = timetables[key_tt].loc[day]
                        if any('(practical)' in str(item) for item in day_schedule if item):
                            has_practical = True
                    if has_practical:
                        continue
                # Check no back-to-back theory and practical for same course
                adjacent_diff_type = False
                for div in session['divs']:
                    key_tt = f"Year{session['year']}_Div{div}"
                    tt = timetables[key_tt]
                    slot_idx = lecture_slot_times.index(slot_time)
                    # Check prev to first
                    if slot_idx > 0:
                        prev_slot = lecture_slot_times[slot_idx - 1]
                        prev_entry = tt.at[day, prev_slot]
                        if prev_entry and session['name'] in prev_entry and session['type'] not in prev_entry:
                            adjacent_diff_type = True
                    # Check next to last
                    last_idx = slot_idx if duration == 1 else slot_idx + 1
                    if last_idx < len(lecture_slot_times) - 1:
                        next_slot = lecture_slot_times[last_idx + 1]
                        next_entry = tt.at[day, next_slot]
                        if next_entry and session['name'] in next_entry and session['type'] not in next_entry:
                            adjacent_diff_type = True
                if adjacent_diff_type:
                    continue
                possible_slots.append((day, slot_time, chosen_room, slot_time2) if duration == 2 else (day, slot_time, chosen_room))
        if not possible_slots:
            print(f"Cannot schedule session: {session}")
            continue
        def get_day_hours(d):
            return len([s for dd, s in faculty_schedule[faculty] if dd == d])
        possible_slots = sorted(possible_slots, key=lambda x: get_day_hours(x[0]))
        if duration == 2:
            day, slot_time1, room, slot_time2 = possible_slots[0]
            faculty_schedule[faculty].append((day, slot_time1))
            faculty_schedule[faculty].append((day, slot_time2))
            room_schedule[room].append((day, slot_time1))
            room_schedule[room].append((day, slot_time2))
            faculty_hours[faculty] += 2
            entry = f"{session['name']} ({session['type']}) - {faculty} - {room}"
            for div in session['divs']:
                key_tt = f"Year{session['year']}_Div{div}"
                timetables[key_tt].at[day, slot_time1] = entry
                timetables[key_tt].at[day, slot_time2] = entry + " (cont.)"
        else:
            day, slot_time, room = possible_slots[0]
            faculty_schedule[faculty].append((day, slot_time))
            room_schedule[room].append((day, slot_time))
            faculty_hours[faculty] += 1
            entry = f"{session['name']} ({session['type']}) - {faculty} - {room}"
            for div in session['divs']:
                key_tt = f"Year{session['year']}_Div{div}"
                timetables[key_tt].at[day, slot_time] = entry
    # --- Fill break slots ---
    break_slots = [s for s in slots if s.startswith('12:15') or s.startswith('15:15')]
    for key in timetables:
        for day in WORKING_DAYS:
            for bs in break_slots:
                if bs.startswith('12:15'):
                    timetables[key].at[day, bs] = "Lunch Break"
                else:
                    timetables[key].at[day, bs] = "Tea Break"
    # --- Export to Excel ---
    temp_dir = tempfile.mkdtemp()
    temp_file = os.path.join(temp_dir, "temp_timetables.xlsx")
    with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
        for key in timetables:
            sheet_name = key.replace(" ", "_")[:31]
            timetables[key].to_excel(writer, sheet_name=sheet_name)
   
    # Merge cells for continued sessions
    wb = load_workbook(temp_file)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        slot_to_col = {}
        for col in range(2, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header in lecture_slot_times:
                slot_to_col[header] = col
        for r in range(2, ws.max_row + 1):
            col = 2
            while col <= ws.max_column:
                cell = ws.cell(r, col)
                if cell.value and ' (cont.)' in str(cell.value):
                    prev_col = col - 1
                    ws.merge_cells(start_row=r, start_column=prev_col, end_row=r, end_column=col)
                    prev_cell = ws.cell(r, prev_col)
                    prev_cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.value = None
                    col += 1  # skip the merged column
                col += 1
    wb.save(temp_file)
    return temp_file

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['GET', 'POST'])
def upload_data():
    if request.method == 'POST':
        # Check if files are uploaded
        if 'course_file' not in request.files or 'faculty_file' not in request.files or 'room_file' not in request.files:
            flash('Please upload all three files: courses, faculty, and rooms.', 'error')
            return redirect(request.url)
       
        course_file = request.files['course_file']
        faculty_file = request.files['faculty_file']
        room_file = request.files['room_file']
       
        if course_file.filename == '' or faculty_file.filename == '' or room_file.filename == '':
            flash('No selected file for one or more inputs.', 'error')
            return redirect(request.url)
       
        # Save uploaded files temporarily
        course_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(course_file.filename))
        faculty_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(faculty_file.filename))
        room_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(room_file.filename))
       
        course_file.save(course_path)
        faculty_file.save(faculty_path)
        room_file.save(room_path)
       
        try:
            # Read the data
            dfCourse = pd.read_csv(course_path)  # Assume CSV; change to pd.read_excel if XLSX
            dfFaculty = pd.read_csv(faculty_path)
            dfRoom = pd.read_csv(room_path)
           
            # Preprocess uploaded data
            dfCourse, dfFaculty, dfRoom = preprocess_data(dfCourse, dfFaculty, dfRoom)
           
            # Clean up uploaded files after reading
            os.remove(course_path)
            os.remove(faculty_path)
            os.remove(room_path)
           
            # Store in session (serialize to JSON)
            session['dfCourse'] = dfCourse.to_json(orient='records')
            session['dfFaculty'] = dfFaculty.to_json(orient='records')
            session['dfRoom'] = dfRoom.to_json(orient='records')
           
            flash('Data uploaded successfully! You can now generate the timetable.', 'success')
            return redirect(url_for('generate_page'))  # Redirect to generate page
        except Exception as e:
            flash(f'Error reading files: {str(e)}', 'error')
            return redirect(request.url)
   
    return render_template('upload.html')

@app.route('/generate_page')
def generate_page():
    # Page to show after upload, with generate button
    return render_template('generate.html')

@app.route('/generate', methods=['POST'])
def generate():
    try:
        # If session data exists, use it; else fallback to default
        if session.get('dfCourse'):
            dfCourse = pd.read_json(session['dfCourse'], orient='records')
            dfFaculty = pd.read_json(session['dfFaculty'], orient='records')
            dfRoom = pd.read_json(session['dfRoom'], orient='records')
        else:
            dfCourse, dfFaculty, dfRoom = get_default_data()
       
        excel_file = generate_timetables(dfCourse, dfFaculty, dfRoom)
        session['timetable_file'] = excel_file
        flash('Timetable generated successfully!', 'success')
        return redirect(url_for('timetable'))
    except Exception as e:
        flash(f'Error generating timetable: {str(e)}', 'error')
        return redirect(url_for('generate_page'))

@app.route('/timetable')
def timetable():
    if 'timetable_file' not in session:
        flash('No timetable generated yet. Please generate one first.', 'error')
        return redirect(url_for('generate_page'))
    
    file_path = session['timetable_file']
    try:
        wb = load_workbook(file_path)
        tables = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append([str(cell) if cell is not None else '' for cell in row])
            tables[sheet_name] = data
        return render_template('timetable.html', tables=tables)
    except Exception as e:
        flash(f'Error loading timetable: {str(e)}', 'error')
        return redirect(url_for('generate_page'))

@app.route('/download_timetable')
def download_timetable():
    if 'timetable_file' not in session:
        flash('No timetable to download. Please generate one first.', 'error')
        return redirect(url_for('generate_page'))
    
    file_path = session['timetable_file']
    return send_file(file_path, as_attachment=True, download_name='Plannable_Timetables.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/demo')
def demo():
    try:
        dfCourse, dfFaculty, dfRoom = get_default_data()
        excel_file = generate_timetables(dfCourse, dfFaculty, dfRoom)
        session['timetable_file'] = excel_file
        flash('Demo timetable generated successfully!', 'success')
        return redirect(url_for('timetable'))
    except Exception as e:
        flash(f'Error generating demo timetable: {str(e)}', 'error')
        return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True, port=5001)